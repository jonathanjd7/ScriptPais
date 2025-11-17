import time
import os
import openpyxl
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.edge.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.microsoft import EdgeChromiumDriverManager
from selenium.common.exceptions import ElementClickInterceptedException, TimeoutException, NoSuchElementException

# Importar configuración
try:
    import config
except ImportError:
    print("ERROR: No se encontró el archivo config.py")
    print("Por favor, copia config.example.py como config.py y completa las credenciales.")
    exit(1)

# Función para eliminar caracteres especiales
def corregir_caracteres(texto):
    if isinstance(texto, str):
        texto = texto.encode('utf-8', 'ignore').decode('utf-8')
        texto = texto.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
        texto = texto.replace('Á', 'A').replace('É', 'E').replace('Í', 'I').replace('Ó', 'O').replace('Ú', 'U')
        texto = texto.replace('ñ', 'n').replace('Ñ', 'N')
    return texto

# Función para crear la URL nueva que se almacena en Excel
def crear_nueva_url(domain, title):
    words = title.split(' ')
    path = ''
    for current, next in zip(words, words[1:] + [None]):
        if next is None:
            path += f"{current}"
        else:
            path += f"{current}-"
    return f"{domain}{corregir_caracteres(path.lower())}/"

# Función para crear la lista de títulos
def crear_titulos(sheet):
    titles = []
    for row in sheet.iter_rows(min_col=1, max_col=1):

        if row[0].value is not None:
            if row[0].value != 'URL title':
                titles.append(row[0].value)
        else:
            break
    return titles

# Configura las opciones del navegador
options = Options()
if config.HEADLESS_MODE:
    options.add_argument("--headless")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

# Usa WebDriver Manager para gestionar el driver automáticamente
# Soporta Edge, Chrome y Firefox según configuración
if config.BROWSER.lower() == 'edge':
    driver = webdriver.Edge(service=Service(EdgeChromiumDriverManager().install()), options=options)
elif config.BROWSER.lower() == 'chrome':
    from webdriver_manager.chrome import ChromeDriverManager
    from selenium.webdriver.chrome.service import Service as ChromeService
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)
elif config.BROWSER.lower() == 'firefox':
    from webdriver_manager.firefox import GeckoDriverManager
    from selenium.webdriver.firefox.service import Service as FirefoxService
    driver = webdriver.Firefox(service=FirefoxService(GeckoDriverManager().install()), options=options)
else:
    raise ValueError(f"Navegador no soportado: {config.BROWSER}. Usa 'edge', 'chrome' o 'firefox'")

# Tamaño de ventana específico para asegurar que los elementos necesarios sean visibles
driver.set_window_size(config.WINDOW_SIZE[0], config.WINDOW_SIZE[1])

# Dirección del fichero Excel donde están los títulos
# Usa ruta relativa al script para portabilidad
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
# Construye la ruta completa desde la ruta relativa del config
excel_path = config.EXCEL_FILE_PATH.replace('./', '')
EXCEL_FILEPATH = os.path.join(SCRIPT_DIR, excel_path)

# URLs desde configuración
url = config.WORDPRESS_LOGIN_URL
url_base = config.WORDPRESS_BASE_PRODUCT_URL
dominio = config.WORDPRESS_PRODUCT_DOMAIN

# Cargar archivos Excel con openpyxl
titulos = openpyxl.load_workbook(EXCEL_FILEPATH)
sheet = titulos.active  # Usar la primera hoja activa

# Crear la lista de títulos
titles = crear_titulos(sheet)

try:
    # Abre la página
    driver.get(url)

    # Espera a que los campos de usuario y contraseña estén visibles
    username_field = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.NAME, 'log'))
    )
    password_field = WebDriverWait(driver, 20).until(
        EC.visibility_of_element_located((By.NAME, 'pwd'))
    )

    # Establece el valor usando JavaScript (desde configuración)
    driver.execute_script("arguments[0].value = arguments[1];", username_field, config.WORDPRESS_USERNAME)
    driver.execute_script("arguments[0].value = arguments[1];", password_field, config.WORDPRESS_PASSWORD)

    # Clic en el botón de inicio de sesión
    login_button = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.ID, 'wp-submit'))
    )
    login_button.click()

    # Espera a que la URL cambie después del primer login
    WebDriverWait(driver, 30).until(EC.url_changes(url))
    
    # Índice de la celda respectiva en el fichero Excel para guardar el siguiente URL
    # Se comienza por la celda 2 para obviar la fila de encabezado
    row_number = 2

    print(f"Se crearán {len(titles)} nuevas URLs.")

    for title in titles:

        try:

            # Nos trasladamos a la URL que sirve de plantilla para duplicar 
            driver.get(url_base)
            WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, "body")))
            print(f"Ingresando a la página de plantilla {url_base}...")

            # Entramos a la página de edición
            edit_product_button = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.LINK_TEXT, 'Editar producto'))
            )
            edit_product_button.click()
            print("Se activó el botón 'Editar producto'")

            # Crear el duplicado
            duplicate_product_button = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="duplicate-action"]/a'))
            )
            duplicate_product_button.click()
            print("Se activó el botón 'Copiar a un nuevo borrador'")

            # Abrir el campo de edición de URL
            edit_slug_button = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="edit-slug-buttons"]/button'))
            )
            edit_slug_button.click()
            print("Se activó el botón 'Editar'")

            # Ingresar el título de la nueva página
            slug_field = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.ID, 'new-post-slug'))
            )
            slug_field.clear()
            slug_field.send_keys(title)
            print(f"Se rellenó el campo de enlace permanente con el título: {title}")

            # Confirmar el cambio hecho
            save_slug_button = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((By.XPATH, '//*[@id="edit-slug-buttons"]/button[1]'))
            )
            save_slug_button.click()
            print("Se activó el botón 'Aceptar'")

            # Publicar los cambios
            publicar_button = WebDriverWait(driver, 30).until(
                EC.presence_of_element_located((By.ID, "publish"))
            )
            driver.execute_script("arguments[0].scrollIntoView(true);", publicar_button)
            time.sleep(5)
            driver.execute_script("arguments[0].click();", publicar_button)
            print("Se activó el botón 'Publicar'")

            # Breve tiempo de espera para asegurar la publicación del URL
            time.sleep(5)
            print(f"URL de {title} creada con éxito.")

            # Ingresar la URL publicada en el fichero Excel
            nueva_url = crear_nueva_url(dominio, title)
            sheet.cell(row=row_number, column=7, value=nueva_url)

            # Guardar cambios de Excel
            titulos.save(EXCEL_FILEPATH)
            print(f"Dirección {nueva_url} ingresada en el fichero Excel {EXCEL_FILEPATH}")
            row_number += 1

        except TimeoutException:
            print(f"Tiempo agotado durante la duplicacion de {title}")
        except Exception as err:
            print(f"Error durante la duplicacion de {title}. Error: {err}")

finally:
    # Espera antes de cerrar el navegador
    time.sleep(20)  
    # Cierra el navegador
    driver.quit()
    print("Navegador cerrado.")